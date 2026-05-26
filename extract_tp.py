import openpyxl
import json
import re

wb = openpyxl.load_workbook('eraport.xlsm', data_only=True)
sheet = wb['EDIT TP']

col_mapping = {
    2: {'kelas': '1', 'semester': 'ganjil'},
    6: {'kelas': '1', 'semester': 'genap'},
    11: {'kelas': '2', 'semester': 'ganjil'},
    15: {'kelas': '2', 'semester': 'genap'},
    20: {'kelas': '3', 'semester': 'ganjil'},
    24: {'kelas': '3', 'semester': 'genap'},
    29: {'kelas': '4', 'semester': 'ganjil'},
    33: {'kelas': '4', 'semester': 'genap'},
    38: {'kelas': '5', 'semester': 'ganjil'},
    42: {'kelas': '5', 'semester': 'genap'},
    47: {'kelas': '6', 'semester': 'ganjil'},
    51: {'kelas': '6', 'semester': 'genap'},
}

extracted_data = []
max_row = sheet.max_row
max_col = 52 # sampai AZ

mapel_blocks = []
# Cari baris-baris di mana nama mapel berada. Biasanya ada di kolom B (index 1) atau judul di baris.
for row_idx in range(1, max_row + 1):
    cell_val = sheet.cell(row=row_idx, column=2).value # Kolom B
    if isinstance(cell_val, str) and len(cell_val) > 4 and cell_val.isupper():
        # Abaikan baris-baris header atau footer
        if "JUMLAH" in cell_val or "KELAS" in cell_val or "SEMESTER" in cell_val or "NO. TP" in cell_val:
            continue
        mapel_blocks.append({'name': cell_val.strip(), 'row': row_idx})

# Perbaiki nama mapel agar sesuai dengan yang di database
def normalize_mapel(name):
    name = name.lower()
    if 'islam' in name: return 'Pendidikan Agama Islam dan Budi Pekerti'
    if 'kristen' in name: return 'Pendidikan Agama Kristen dan Budi Pekerti'
    if 'pancasila' in name: return 'Pendidikan Pancasila'
    if 'indonesia' in name: return 'Bahasa Indonesia'
    if 'matematika' in name: return 'Matematika'
    if 'ipas' in name or 'alam dan sosial' in name: return 'Ilmu Pengetahuan Alam dan Sosial (IPAS)'
    if 'musik' in name: return 'Seni Musik'
    if 'rupa' in name: return 'Seni Rupa'
    if 'tari' in name: return 'Seni Tari'
    if 'teater' in name: return 'Seni Teater'
    if 'jasmani' in name or 'pjok' in name: return 'Pendidikan Jasmani Olahraga dan Kesehatan (PJOK)'
    if 'inggris' in name: return 'Bahasa Inggris'
    if 'jawa' in name: return 'Bahasa Jawa'
    if 'arab' in name: return 'Bahasa Arab'
    if 'mulok' in name: return name.title()
    return name.title()

for block in mapel_blocks:
    mapel_name = normalize_mapel(block['name'])
    start_tp_row = block['row'] + 4
    
    for col_idx, meta in col_mapping.items():
        tp_number = 1
        for offset in range(15): # Ada maksimal sekitar 10-15 TP
            cell_val = sheet.cell(row=start_tp_row + offset, column=col_idx + 1).value
            
            if cell_val and isinstance(cell_val, str) and len(cell_val.strip()) > 5:
                desc = cell_val.strip()
                # Abaikan teks header/footer atau sel kosong semu
                if desc == '-' or 'JUMLAH TP' in desc.upper() or desc.upper() == 'TP' or desc.isdigit():
                    continue
                    
                extracted_data.append({
                    'mapel_name': mapel_name,
                    'kelas': meta['kelas'],
                    'semester': meta['semester'],
                    'nomor_tp': tp_number,
                    'deskripsi': desc
                })
                tp_number += 1

with open('database/seeders/targets_data.json', 'w', encoding='utf-8') as f:
    json.dump(extracted_data, f, ensure_ascii=False, indent=2)

print(f"Berhasil mengekstrak {len(extracted_data)} TP/CP.")
