import openpyxl

wb = openpyxl.load_workbook('eraport.xlsm', data_only=True)

print("=== SHEET: Legger ===")
sheet_legger = wb['Legger']
# Mari kita cetak beberapa baris dan kolom (misal baris 7 sampai 10, kolom A sampai Z)
for row in sheet_legger.iter_rows(min_row=6, max_row=10, min_col=1, max_col=35, values_only=True):
    # filter None or empty string to make it readable
    print([str(val)[:20] if val is not None else "" for val in row])

print("\n=== SHEET: EKSPOR (Format Rapor) ===")
sheet_ekspor = wb['EKSPOR']
# Mari kita lihat beberapa area penting (misal header dan isi rapor)
for row in sheet_ekspor.iter_rows(min_row=1, max_row=30, min_col=1, max_col=10, values_only=True):
    print([str(val)[:30] if val is not None else "" for val in row if val is not None])
